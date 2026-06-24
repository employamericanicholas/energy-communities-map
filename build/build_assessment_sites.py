"""Build docs/data/assessment_sites.geojson from the Employ America
"Site Assessment By Company" workbook (Potential Sites + Potential Firms).

Coordinates, in priority order:
  1. RESEARCHED  - precise coordinates looked up for discrete facilities
     (geothermal plants, pumped-storage projects, CCS, a few nuclear sites).
  2. reuse from docs/data/nuclear.geojson - for sites already mapped there
     (the bulk of the nuclear plants), matched by alias-normalized name.
  3. county centroid - fallback for multi-county lease areas (e.g. Fervo
     Pipeline blocks), early-feasibility sites, and projects with no
     published coordinate. These carry a loc_note.

Also attaches Employ America's point-based unemployment margins (area rate
minus the national-average threshold, in percentage points) for the two
notice vintages: margin_2531 (2024 "Last Guidance") and margin_2639 (2025
"Current Method"), plus the "+ Micro & CSA" alternatives. Positive = the
area's unemployment is at/above the national average (meets the criterion).
"""
import json, os, math, re
import openpyxl
from shapely.geometry import shape, Point
from shapely import STRtree

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(ROOT, "docs", "data")
XLSX = os.path.join(ROOT, "build", "site_assessment.xlsx")

COUNTY_FIX = {
    "Southeastern Connecticut": ("New London County", "CT"),
    "Southeastern Connecticut Planning Region, CT": ("New London County", "CT"),
}

# Precise coordinates (lon, lat[, expected_state]) keyed by normalized site name.
# Sourced from Wikipedia infoboxes / Global Energy Monitor / official pages.
RESEARCHED = {
    "mcginnesshills": (-116.9117, 39.5892),
    "hebercomplex": (-115.5180, 32.7146, "CA"),  # Heber, Imperial County CA only
    "beowawe": (-116.6175, 40.5547),
    "tuscarora": (-116.1506, 41.4673),
    "covefort": (-112.5850, 38.5750),
    "puna": (-154.8888, 19.4785),
    "thegeysers": (-122.7558, 38.7906),
    "saltonsea": (-115.6200, 33.2000),
    "newberry": (-121.2290, 43.7220),
    "bathcountypumpedstorage": (-79.8194, 38.2306),
    "ludingtonpumpedstorage": (-86.4453, 43.8936),
    "raccoonmountainpumpedstorage": (-85.3967, 35.0483),
    "helmspumpedstorage": (-118.9647, 37.0369),
    "eaglemountainpumpedstorage": (-115.4900, 33.8700),
    "gordonbuttepumpedstorage": (-110.3333, 46.4167),
    "calpinebaytownccsproject": (-94.9019, 29.7731),
    "broadwingenergycenter": (-88.8810, 39.8320),
    "levy": (-82.6217, 29.0733),
    "donaldccook": (-86.5653, 41.9756),
    "bellefonte": (-85.9292, 34.7086),
    "hermes": (-84.3960, 35.9210),
    "victoriastation": (-97.0306, 28.6144),
}
# substring -> (lon, lat) for long compound site names
RESEARCHED_SUB = {
    "dixievalley": (-117.8557, 39.9663),
    "castaic": (-118.6566, 34.5873),
}


def normname(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


# Workbook site name -> nuclear.geojson feature name (alias) for reuse.
ALIAS = {
    "cyrstalriver": "crystalriver", "tmi": "threemileisland", "vcsummer": "virgilcsummer",
    "southtexasprojectabwr": "southtexas", "salemhopecreek": "salem", "bellbendsusquehanna": "susquehanna",
    "fermiunit3": "enricofermi", "coopernuclearstation": "cooper", "terrapowernatriumkemmerer": "terrapower",
    "seadrift": "longmott", "greenfield": "stewartcounty",
    "hatch": "edwinihatch", "farley": "josephmfarley", "ginna": "reginna",
}


def etype_of(source):
    s = (source or "").lower()
    if "geotherm" in s: return "Geothermal"
    if "nuclear" in s: return "Nuclear"
    if "ccs" in s: return "CCS"
    if "storage" in s: return "Storage"
    return "Other"


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    firm_src = {}
    for r in list(wb["Potential Firms"].iter_rows(values_only=True))[1:]:
        if r[0] and r[1]:
            firm_src[str(r[0]).strip()] = str(r[1]).strip()

    def company_etype(comp):
        if comp in firm_src:
            return etype_of(firm_src[comp])
        for k, v in firm_src.items():
            if k.lower() in comp.lower() or comp.lower() in k.lower():
                return etype_of(v)
        return "Other"

    def site_etype(comp, site):
        s = (site or "").lower()
        if "ccs" in s: return "CCS"
        if "pumped" in s or "storage" in s: return "Storage"
        return company_etype(comp)

    # County centroid lookup + precise coords reused from nuclear.geojson
    counties = json.load(open(os.path.join(D, "counties.geojson")))
    cidx = {}
    for f in counties["features"]:
        p = f["properties"]
        rp = shape(f["geometry"]).representative_point()
        cidx[(p["NAMELSAD"].lower(), p["STUSPS"])] = (round(rp.x, 4), round(rp.y, 4), p["GEOID"], p["STATE_NAME"])

    nuc = json.load(open(os.path.join(D, "nuclear.geojson")))
    nuc_by_norm = {}
    for f in nuc["features"]:
        nuc_by_norm[normname(f["properties"]["name"])] = (
            f["geometry"]["coordinates"], bool(f["properties"].get("loc_note")))

    def resolve_county(cty):
        if cty in COUNTY_FIX:
            name, st = COUNTY_FIX[cty]
        elif "," in cty:
            left, st = cty.rsplit(",", 1); st = st.strip(); name = left.strip()
        else:
            return None, None
        name2 = name if name.lower().endswith("county") else name + " County"
        for key in [(name.lower(), st), (name2.lower(), st)]:
            if key in cidx:
                return cidx[key], st
        return None, st

    def reuse_nuclear(site):
        n = normname(site)
        n = ALIAS.get(n, n)
        if n in nuc_by_norm:
            return nuc_by_norm[n]
        for k, v in nuc_by_norm.items():     # prefix match (either direction)
            if n and len(n) >= 4 and (k.startswith(n) or n.startswith(k)):
                return v
        return None

    rows = list(wb["Potential Sites"].iter_rows(values_only=True))[2:]
    feats, per_county = [], {}
    for r in rows:
        if not (r[1] or r[2]):
            continue
        comp = str(r[0]).strip() if r[0] else ""
        cty = str(r[1]).strip() if r[1] else ""
        site = str(r[2]).strip() if r[2] else ""
        cinfo, st = resolve_county(cty)
        nm = normname(site)

        lon = lat = None; precise = False; note = None
        # 1. researched precise
        if nm in RESEARCHED:
            val = RESEARCHED[nm]
            if len(val) == 2 or val[2] == st:
                lon, lat = val[0], val[1]; precise = True
        if lon is None:
            for sub, c in RESEARCHED_SUB.items():
                if sub in nm:
                    lon, lat = c; precise = True; break
        # 2. reuse from nuclear.geojson
        if lon is None:
            ru = reuse_nuclear(site)
            if ru:
                (lon, lat), approx = ru
                precise = not approx
                if approx:
                    note = "Approximate location (no surveyed coordinate published)."
        # 3. county centroid
        if lon is None:
            if not cinfo:
                continue
            lon, lat = cinfo[0], cinfo[1]
            n = per_county.get(cinfo[2], 0); per_county[cinfo[2]] = n + 1
            if n:
                ang = n * 2.399963
                lon += 0.06 * math.cos(ang); lat += 0.06 * math.sin(ang)
            note = "Mapped at the county centroid — energy-community eligibility is determined at the county / MSA level, not the exact site point."

        state = (cinfo[2] if False else (cinfo[3] if cinfo else st))
        geoid = cinfo[2] if cinfo else ""
        name = site if site and site != "(Greenfield)" else f"{comp} (greenfield)"
        et = site_etype(comp, site)

        def num(x):
            return round(float(x), 2) if isinstance(x, (int, float)) else None
        props = {
            "name": name, "owner": comp, "etype": et, "status": et,
            "county": None, "state": state, "GEOID": geoid, "site_county": cty,
            "margin_2531": num(r[3]), "margin_2639": num(r[4]),
            "margin_2531_mc": num(r[5]), "margin_2639_mc": num(r[6]),
        }
        if note:
            props["loc_note"] = note
        feats.append({"type": "Feature", "properties": props,
                      "geometry": {"type": "Point", "coordinates": [round(lon, 4), round(lat, 4)]}})

    # ---- IRS eligibility by point-in-polygon ----
    def index_layer(path):
        fs = json.load(open(path))["features"]
        geoms = [shape(f["geometry"]) for f in fs]
        return STRtree(geoms), geoms, fs

    def locate(pt, tree, geoms, fs):
        for i in tree.query(pt):
            if geoms[i].contains(pt):
                return fs[i]["properties"]
        return None

    ffe = index_layer(os.path.join(D, "ffe_counties.geojson"))
    coal = index_layer(os.path.join(D, "coal_tracts.geojson"))
    cty_layer = index_layer(os.path.join(D, "counties.geojson"))
    for f in feats:
        pt = Point(*f["geometry"]["coordinates"])
        cc = locate(pt, *cty_layer)
        f["properties"]["county"] = cc.get("NAMELSAD", "") if cc else ""
        fc = locate(pt, *ffe)
        f["properties"]["ffe"] = bool(fc and (fc.get("do_2531") or fc.get("do_2639") or fc.get("may")))
        f["properties"]["ffe_unemp_2531"] = bool(fc and fc.get("do_2531"))
        f["properties"]["ffe_unemp_2639"] = bool(fc and fc.get("do_2639"))
        ct = locate(pt, *coal)
        f["properties"]["coal_2639"] = ct is not None
        f["properties"]["coal_2531"] = bool(ct and ct.get("since") != "2026-39")

    json.dump({"type": "FeatureCollection", "features": feats},
              open(os.path.join(D, "assessment_sites.geojson"), "w"))
    from collections import Counter
    approx = sum(1 for f in feats if f["properties"].get("loc_note"))
    print(f"assessment sites: {len(feats)}  precise: {len(feats)-approx}  approx/centroid: {approx}")
    print("by type:", dict(Counter(f["properties"]["etype"] for f in feats)))


if __name__ == "__main__":
    main()
