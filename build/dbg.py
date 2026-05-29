import openpyxl
MAYX="Counties and Census Tracts Eligable for Coal and FFE/Eligible FFE Counties that MAY meet Unemployment Threshold/n-25-31-appendix-2.xlsx"
wb=openpyxl.load_workbook(MAYX, read_only=True, data_only=True); ws=wb.active
none0=0; ok=0; tot=0
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i<2: continue
    tot+=1
    if r[0] is None: none0+=1
    if i in (2,3,50,100,200,300,400,440,443,444,445): print(i, r)
print("total data rows:",tot,"col0 None:",none0)
wb.close()
DO="Counties and Census Tracts Eligable for Coal and FFE/Eligible FFE Counties that DO meet Unemployment/MSA_FFEunemployment.xlsx"
wb=openpyxl.load_workbook(DO, read_only=True, data_only=True); ws=wb.active
tot=0;none0=0
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i<3: continue
    tot+=1
    if r[0] is None: none0+=1
print("DO total data rows:",tot,"col0 None:",none0)
wb.close()
